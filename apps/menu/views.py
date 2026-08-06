import csv

from openpyxl import load_workbook
from openpyxl import Workbook

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.urls import reverse_lazy
from django.views.generic import (
    ListView,
    CreateView,
    UpdateView,
    DeleteView,
)

from apps.accounts.decorators import permission_required

from .forms import (
    MenuCategoryForm,
    MenuItemForm,
)

from .models import (
    MenuCategory,
    MenuItem,
)


# ==========================================================
# MENU CATEGORY
# ==========================================================

class MenuCategoryListView(ListView):

    model = MenuCategory

    template_name = "menu/category_list.html"

    context_object_name = "categories"


class MenuCategoryCreateView(CreateView):

    model = MenuCategory

    form_class = MenuCategoryForm

    template_name = "menu/category_form.html"

    success_url = reverse_lazy("menu:category-list")

    def form_valid(self, form):

        messages.success(
            self.request,
            "Category added successfully."
        )

        return super().form_valid(form)


class MenuCategoryUpdateView(UpdateView):

    model = MenuCategory

    form_class = MenuCategoryForm

    template_name = "menu/category_form.html"

    success_url = reverse_lazy("menu:category-list")

    def form_valid(self, form):

        messages.success(
            self.request,
            "Category updated successfully."
        )

        return super().form_valid(form)


class MenuCategoryDeleteView(DeleteView):

    model = MenuCategory

    template_name = "menu/category_confirm_delete.html"

    success_url = reverse_lazy("menu:category-list")

    def delete(self, request, *args, **kwargs):

        messages.success(
            request,
            "Category deleted successfully."
        )

        return super().delete(request, *args, **kwargs)


# ==========================================================
# MENU ITEMS
# ==========================================================

class MenuItemListView(ListView):

    model = MenuItem

    template_name = "menu/item_list.html"

    context_object_name = "items"

    paginate_by = 20


    def get_queryset(self):

        queryset = MenuItem.objects.select_related(
            "category"
        ).order_by("name")

        search = self.request.GET.get(
            "search",
            ""
        )

        category = self.request.GET.get(
            "category",
            ""
        )

        veg = self.request.GET.get(
            "veg",
            ""
        )

        status = self.request.GET.get(
            "status",
            ""
        )

        if search:

            queryset = queryset.filter(
                name__icontains=search
            )

        if category:

            queryset = queryset.filter(
                category_id=category
            )

        if veg == "veg":

            queryset = queryset.filter(
                is_veg=True
            )

        elif veg == "nonveg":

            queryset = queryset.filter(
                is_veg=False
            )

        if status == "available":

            queryset = queryset.filter(
                is_available=True
            )

        elif status == "unavailable":

            queryset = queryset.filter(
                is_available=False
            )

        return queryset


    def get_context_data(self, **kwargs):

        context = super().get_context_data(**kwargs)

        context["categories"] = MenuCategory.objects.all()

        return context
    
class MenuItemCreateView(CreateView):

    model = MenuItem

    form_class = MenuItemForm

    template_name = "menu/item_form.html"

    success_url = reverse_lazy("menu:item-list")

    def form_valid(self, form):

        messages.success(
            self.request,
            "Menu item added successfully."
        )

        return super().form_valid(form)


class MenuItemUpdateView(UpdateView):

    model = MenuItem

    form_class = MenuItemForm

    template_name = "menu/item_form.html"

    success_url = reverse_lazy("menu:item-list")

    def form_valid(self, form):

        messages.success(
            self.request,
            "Menu item updated successfully."
        )

        return super().form_valid(form)


class MenuItemDeleteView(DeleteView):

    model = MenuItem

    template_name = "menu/item_confirm_delete.html"

    success_url = reverse_lazy("menu:item-list")

    def form_valid(self, form):

        messages.success(
            self.request,
            "Menu item deleted successfully."
        )

        return super().form_valid(form)


# ==========================================================
# IMPORT MENU ITEMS
# ==========================================================

@login_required
@permission_required("menu.add")
def import_menu_items(request):

    if request.method == "POST":

        file = request.FILES.get("file")

        if not file:

            messages.error(
                request,
                "Please select a file."
            )

            return redirect("menu:item-list")

        try:

            # ---------------- CSV ----------------

            if file.name.endswith(".csv"):

                decoded = file.read().decode("utf-8").splitlines()

                reader = csv.DictReader(decoded)

                for row in reader:

                    category, _ = MenuCategory.objects.get_or_create(
                        name=row["Category"]
                    )

                    MenuItem.objects.create(
                        category=category,
                        name=row["Name"],
                        price=float(row["Price"]),
                        preparation_time=int(
                            row["Preparation Time"]
                        ),
                        is_veg=row["Veg"].lower() == "yes",
                        is_available=True,
                    )

            # ---------------- XLSX ----------------

            elif file.name.endswith(".xlsx"):

                workbook = load_workbook(file)

                sheet = workbook.active

                for row in sheet.iter_rows(
                    min_row=2,
                    values_only=True,
                ):

                    category, _ = MenuCategory.objects.get_or_create(
                        name=row[1]
                    )

                    MenuItem.objects.create(
                        name=row[0],
                        category=category,
                        price=row[2],
                        preparation_time=row[3],
                        is_veg=str(row[4]).lower() == "yes",
                        is_available=True,
                    )

            else:

                messages.error(
                    request,
                    "Only CSV and XLSX files are supported."
                )

                return redirect("menu:item-list")

            messages.success(
                request,
                "Menu imported successfully."
            )

            return redirect("menu:item-list")

        except Exception as e:

            messages.error(
                request,
                f"Import failed: {e}"
            )

            return redirect("menu:item-list")

    return render(
        request,
        "menu/import_items.html",
    )


# ==========================================================
# EXPORT MENU ITEMS
# ==========================================================

@login_required
@permission_required("menu.view")
def export_menu_items(request):

    export_type = request.GET.get(
        "type",
        "csv",
    )

    items = MenuItem.objects.select_related(
        "category"
    )

    # ======================================================
    # CSV
    # ======================================================

    if export_type == "csv":

        response = HttpResponse(
            content_type="text/csv"
        )

        response[
            "Content-Disposition"
        ] = 'attachment; filename="menu_items.csv"'

        writer = csv.writer(response)

        writer.writerow([
            "Name",
            "Category",
            "Price",
            "Preparation Time",
            "Veg",
        ])

        for item in items:

            writer.writerow([
                item.name,
                item.category.name,
                item.price,
                item.preparation_time,
                "Yes" if item.is_veg else "No",
            ])

        return response

    # ======================================================
    # XLSX
    # ======================================================

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "Menu Items"

    sheet.append([
        "Name",
        "Category",
        "Price",
        "Preparation Time",
        "Veg",
    ])

    for item in items:

        sheet.append([
            item.name,
            item.category.name,
            item.price,
            item.preparation_time,
            "Yes" if item.is_veg else "No",
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response[
        "Content-Disposition"
    ] = 'attachment; filename="menu_items.xlsx"'

    workbook.save(response)

    return response