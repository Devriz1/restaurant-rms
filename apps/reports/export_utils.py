import csv
from decimal import Decimal
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.template.loader import render_to_string
from xhtml2pdf import pisa


CURRENCY_SYMBOL = "₹"


def _get_value(obj, field_path):
    parts = field_path.split(".")
    current = obj
    for part in parts:
        if isinstance(current, dict):
            current = current.get(part, "")
        elif current is None:
            return ""
        else:
            current = getattr(current, part, "")
    if isinstance(current, Decimal):
        return float(current)
    if isinstance(current, datetime):
        return current.strftime("%d %b %Y %I:%M %p")
    if isinstance(current, date):
        return current.strftime("%d %b %Y")
    return current if current is not None else ""


def _format_row(row, columns):
    return [_get_value(row, field) for field, _ in columns]


def export_csv(data, columns, filename, summary_cards=None, symbol="₹"):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)

    if summary_cards:
        writer.writerow(["Summary"])
        for card in summary_cards:
            writer.writerow([card.get("title", ""), card.get("value", "")])
        writer.writerow([])

    writer.writerow([display for _, display in columns])
    for row in data:
        writer.writerow(_format_row(row, columns))

    return response


def export_excel(data, columns, filename, summary_cards=None, symbol="₹"):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Report"

    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True, color="1F4E79")
    summary_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    row_idx = 1

    if summary_cards:
        for card in summary_cards:
            sheet.cell(row=row_idx, column=1, value=card.get("title", ""))
            sheet.cell(row=row_idx, column=2, value=card.get("value", ""))
            sheet.cell(row=row_idx, column=1).fill = summary_fill
            sheet.cell(row=row_idx, column=2).fill = summary_fill
            row_idx += 1
        row_idx += 1

    for col_idx, (_, display) in enumerate(columns, start=1):
        cell = sheet.cell(row=row_idx, column=col_idx, value=display)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in data:
        row_idx += 1
        for col_idx, value in enumerate(_format_row(row, columns), start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)

    for col_idx in range(1, len(columns) + 1):
        sheet.column_dimensions[get_column_letter(col_idx)].width = 18

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


def export_pdf(data, columns, filename, title, summary_cards=None, symbol="₹"):
    formatted_data = []
    for row in data:
        formatted_row = []
        for value in _format_row(row, columns):
            if isinstance(value, (int, float)) or (isinstance(value, str) and value.replace('.', '', 1).isdigit()):
                formatted_row.append(f"{symbol} {float(value):.2f}")
            else:
                formatted_row.append(value)
        formatted_data.append(formatted_row)

    html_string = render_to_string(
        "reports/pdf_export.html",
        {
            "title": title,
            "columns": columns,
            "data": formatted_data,
            "summary_cards": summary_cards,
            "currency_symbol": symbol,
        },
    )

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    pisa_status = pisa.CreatePDF(html_string, dest=response)
    if pisa_status.err:
        return HttpResponse("Error generating PDF", status=500)

    return response
